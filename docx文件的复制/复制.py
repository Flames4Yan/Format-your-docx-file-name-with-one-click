import shutil
import os
from shutil import which, copyfile

import openpyxl
def copyfi(orindic_path,copydic_path):
    which_type = input("最后后缀:")
    if choose==1:
        key=input("实验几:")
        work_book = openpyxl.load_workbook(f'{orindic_path}/Java实验课名单.xlsx')
        orindic_path=f'{orindic_path}/Java-{key}'

    elif choose==2:
        key = input("实验几:")
        work_book = openpyxl.load_workbook(f'{orindic_path}/计组实验课名单.xlsx')
        orindic_path = f'{orindic_path}/计算机-{key}'

    sheet_name = work_book.sheetnames
    sheet = work_book[sheet_name[0]]
    os.mkdir(f'{copydic_path}/2024-2025-1-{which_type}')# 选择哪个工作表
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        st_code = row[1]
        # 选取遍历目录的
        # 匹配字符串的部分
        part_name = name
        matching_file = []  # list
        for filename in os.listdir(orindic_path):
            if part_name in filename:
                matching_file = filename
                break
        orig_flie_path = f'{orindic_path}/{matching_file}'
        new_flie_path = f'{copydic_path}/2024-2025-1-{which_type}/{st_code}-{name}-{which_type}.docx'
        shutil.copy(orig_flie_path, new_flie_path)
        # 目前测试成功


which_class=0
choose=int(input("实验课为:(1-java,2-计算机组成原理):"))
if choose==1:
    copyfi('./Orin-Java','./Copy-Java')
elif choose==2:
    copyfi('./Orin-计组','./Copy-计组')
#目前测试成功




